"""
Appends 5 new city tabs to asian_hair_barber_seed_list.xlsx: London, Sydney,
Toronto, Amsterdam, Madrid. Chosen as cities with a strong Asian population
and/or Asian travel draw, researched via web search (Reddit/Yelp/Google review
snippets/business sites) rather than a direct Google Places API pull like the
original six cities — so provenance is noted per-row in the Source column, and
address/rating/review-count fields are left blank where a search snippet
didn't surface them (no fabricated business data).

Run: python3 scripts/append_new_cities.py
"""
import openpyxl

FILE = "asian_hair_barber_seed_list.xlsx"

HEADER = ["City", "Neighborhood", "Business", "Address", "Phone", "Google rating",
          "Review count", "Signal strength", "Signal notes (paraphrased)", "Source",
          "Google Maps link"]

ROWS = {
    "London": [
        ("London", "New Malden (Korean high street)", "The Fountain Hair (Korean Hair Salon)",
         "122 Malden Road, New Malden", None, None, None, "High",
         "Salon explicitly described as working with all hair types 'from European to Asian, with thorough understanding of them'; stylist Anna specifically praised for attention to detail.",
         "Web search (business site + Fresha listing)", None),
        ("London", "New Malden (Korean high street)", "TOVI Hair Salon",
         "Park Road, New Malden", None, None, None, "Medium",
         "30 years' experience, offers 'magic straight' treatments (chemical straightening) at affordable prices; located on the same Korean-community high street as other strong leads, but no explicit Asian-hair review quote found yet.",
         "Web search (business site)", None),
        ("London", "New Malden (Korean high street)", "Bokko Korean Hair Salon",
         "New Malden", None, None, None, "Medium",
         "Korean-branded salon on the New Malden high street with strong generic service praise; no explicit Asian-hair quote found yet.",
         "Web search (business site)", None),
        ("London", "West London", "F4Fade",
         None, None, None, None, "Medium",
         "Specializes in Afro, European, and Asian hair; 15+ years in the west London community, with some celebrity clientele mentioned when passing through London.",
         "Web search (press/directory mentions)", None),
    ],
    "Sydney": [
        ("Sydney", "Eastwood (Koreatown)", "DNK Hair",
         "100 Rowe St, Eastwood, NSW", "(02) 9874 4448", None, None, "High",
         "Korean hair salon with 30+ years of heritage from Itaewon, Seoul; well known locally. One review specifically described dryness and manageability problems after a chemical straightening treatment — recommend for cuts, caution on chemical/straightening services specifically.",
         "Web search (business site + Yelp/review aggregator snippets)", None),
        ("Sydney", "Haymarket / multiple", "Kim Sun Young Hair",
         "764 George St, Haymarket, NSW (also World Square, Strathfield, Liverpool St)", None, 4.9, 2541, "High",
         "Korean hair brand dating to 1956 Seoul; Haymarket location rated 4.9★ from 2,541 reviews. Reviews are strong overall but several specifically flag inconsistent perm longevity — recommend for cuts/styling, caution on perm service specifically.",
         "Web search (Fresha + Yelp listing)", None),
    ],
    "Toronto": [
        ("Toronto", "Pacific Mall, Markham", "Artis Hair (Pacific Mall, unit B69)",
         "Pacific Mall, Markham, ON", None, None, None, "None",
         "Located inside Pacific Mall, North America's largest indoor Asian shopping mall, alongside many other Chinese-run hair salons — strong community context, but no business-specific hair-type review signal found yet.",
         "Web search (mall directory)", None),
        ("Toronto", "Pacific Mall, Markham", "G.Z. Hair Salon (Pacific Mall, unit F31)",
         "Pacific Mall, Markham, ON", None, None, None, "None",
         "Same mall / same community context as Artis Hair; no business-specific hair-type review signal found yet.",
         "Web search (mall directory)", None),
        ("Toronto", "Downtown Chinatown", "Unnamed Chinatown Centre hairdressers",
         "222 Spadina Ave (Chinatown Centre), Toronto, ON", None, None, None, "None",
         "General mentions of 'several Chinese hairdressers' inside the Chinatown Centre building; no specific business name or review with explicit hair-type signal surfaced.",
         "Web search (forum mentions)", None),
    ],
    "Amsterdam": [
        ("Amsterdam", "Oud-West", "MensPlace Barbershop (barber: Ogun)",
         "Staringplein 20HS, 1054 VL Amsterdam", None, 4.9, 669, "High",
         "Barber Ogun, cutting since 2016, explicitly described as experienced across 'every hair type and texture: from straight Asian hair to coily Afro hair.' Multiple direct customer quotes praising consistency and fit to request.",
         "Web search (business site + Tripadvisor)", None),
        ("Amsterdam", "City center", "ASSORT AMSTERDAM",
         None, None, None, None, "Medium",
         "Branded specifically as a Japanese hair salon with a bilingual team; branding-level signal only, no specific customer quote about Asian hair confirmed yet.",
         "Web search (business site)", None),
        ("Amsterdam", "Amsterdam / Rotterdam", "Hair Studio Picnic",
         None, None, None, None, "Medium",
         "Japanese hair salon since 2015 offering Japanese straightening and digital perm; branding-level signal only, no specific customer quote confirmed yet.",
         "Web search (business site)", None),
    ],
    "Madrid": [
        ("Madrid", "Usera (Madrid's Chinatown)", "HAIRONE Salon de peluqueria y Barberia (华人美发店)",
         "Usera, Madrid", None, None, None, "Medium",
         "Chinese-run salon (name translates to 'Chinese people's hair salon') in Usera; reviews describe quality haircuts and professional service, though none explicitly name Asian hair texture.",
         "Web search (Booksy listing)", None),
        ("Madrid", "Usera (Madrid's Chinatown)", "Dharma Peluqueros",
         "Calle Marcelo de Usera, Madrid", None, None, None, "None",
         "General salon located in the Usera neighborhood; no explicit Asian-hair review signal found yet.",
         "Web search (directory listing)", None),
    ],
}


def main():
    wb = openpyxl.load_workbook(FILE)
    for city, rows in ROWS.items():
        if city in wb.sheetnames:
            del wb[city]
        ws = wb.create_sheet(city)
        ws.append(HEADER)
        for row in rows:
            ws.append(list(row))
    wb.save(FILE)
    print(f"Added/updated tabs: {list(ROWS.keys())}")


if __name__ == "__main__":
    main()
